"""
Excelパーサー

2つのデータソースから都道府県別清酒データを抽出する:
1. 時系列Excel (jikeiretsu_13.xlsx) — 販売（消費）数量 1963-2023
2. e-Stat統計年報 表0802 — 製成数量（年度別Excel）
"""

import re
from pathlib import Path

import pandas as pd

BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = BASE_DIR / "data" / "raw"
PROCESSED_DIR = BASE_DIR / "data" / "processed"

# 47都道府県の正規化リスト
PREFECTURES = [
    "北海道", "青森", "岩手", "宮城", "秋田", "山形", "福島",
    "茨城", "栃木", "群馬", "埼玉", "千葉", "東京", "神奈川",
    "新潟", "富山", "石川", "福井", "山梨", "長野",
    "岐阜", "静岡", "愛知", "三重",
    "滋賀", "京都", "大阪", "兵庫", "奈良", "和歌山",
    "鳥取", "島根", "岡山", "広島", "山口",
    "徳島", "香川", "愛媛", "高知",
    "福岡", "佐賀", "長崎", "熊本", "大分", "宮崎", "鹿児島", "沖縄",
]


SPECIAL_MAPPINGS = {
    "札幌局計": "北海道",  # 札幌国税局 = 北海道のみ
}


def normalize_pref_name(name):
    """都道府県名を正規化する。全角スペース除去、県/府/都を除去して照合。"""
    if not name or not isinstance(name, str):
        return None
    name = name.strip().replace("\u3000", "").replace(" ", "")

    # 特殊マッピング
    if name in SPECIAL_MAPPINGS:
        return SPECIAL_MAPPINGS[name]

    # 「県」「府」「都」を除去
    name_stripped = re.sub(r"[県府都]$", "", name)
    for pref in PREFECTURES:
        # 「京都」「大阪」は県/府を除去せず完全一致 or 付きで一致
        if pref in ("京都", "大阪"):
            if name == pref or name_stripped == pref:
                return pref
        elif pref == "北海道":
            if name == pref:
                return pref
        elif pref == "東京":
            if name == pref or name_stripped == pref:
                return pref
        else:
            pref_short = re.sub(r"[県府都]$", "", pref)
            if name == pref or name_stripped == pref_short or name == pref_short:
                return pref
    return None


def parse_jikeiretsu():
    """時系列Excel (13.xlsx) から都道府県別販売（消費）数量パネルを抽出。"""
    print("=== 時系列Excel パース ===")
    path = RAW_DIR / "nenpo" / "jikeiretsu_13.xlsx"

    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["２\u3000清酒"]

    # 年度ヘッダを取得（row 4, col 3以降）
    # 全角数字→半角変換テーブル
    zen2han = str.maketrans("０１２３４５６７８９", "0123456789")

    years = []
    era_base = None
    for col in range(3, ws.max_column + 1):
        v = ws.cell(row=4, column=col).value
        if v is None:
            continue
        v_str = str(v).strip().translate(zen2han)

        if "昭和" in v_str:
            era_base = 1925  # 昭和N年 = 1925 + N
            year_num = int(re.search(r"(\d+)", v_str).group(1))
            years.append((col, era_base + year_num))
        elif "平成" in v_str:
            era_base = 1988  # 平成N年 = 1988 + N
            year_num = int(re.search(r"(\d+)", v_str.replace("元", "1")).group(1))
            years.append((col, era_base + year_num))
        elif "令和" in v_str:
            era_base = 2018  # 令和N年 = 2018 + N
            year_num = int(re.search(r"(\d+)", v_str.replace("元", "1")).group(1))
            years.append((col, era_base + year_num))
        elif "都道府県" in v_str:
            continue
        else:
            # 数字のみの場合は前の元号の続き
            try:
                year_num = int(re.search(r"(\d+)", v_str).group(1))
                if era_base:
                    years.append((col, era_base + year_num))
            except (ValueError, AttributeError):
                pass

    # 重複年度の補正: 連続する年度が期待されるため、
    # 重複がある場合は前年+1で補正（元データの誤植対策）
    corrected_years = []
    for i, (col, year) in enumerate(years):
        if i > 0 and year == corrected_years[-1][1]:
            corrected_year = corrected_years[-1][1] + 1
            print(f"  ⚠ 年度重複を補正: col {col} {year} → {corrected_year}")
            corrected_years.append((col, corrected_year))
        else:
            corrected_years.append((col, year))
    years = corrected_years

    print(f"  年度数: {len(years)} ({years[0][1]}〜{years[-1][1]})")

    # 都道府県データを抽出
    records = []
    for row_idx in range(7, ws.max_row + 1):
        # col 1 = 国税局計、col 2 = 都道府県名
        c1 = ws.cell(row=row_idx, column=1).value
        c2 = ws.cell(row=row_idx, column=2).value

        name = c2 or c1
        pref = normalize_pref_name(str(name)) if name else None
        if not pref:
            continue

        for col, year in years:
            val = ws.cell(row=row_idx, column=col).value
            if val is None or val == "-" or val == "X":
                quantity = None
            else:
                try:
                    quantity = float(val)
                except (ValueError, TypeError):
                    quantity = None
            records.append({
                "year": year,
                "prefecture": pref,
                "sales_quantity_kl": quantity,
            })

    df = pd.DataFrame(records)
    print(f"  レコード数: {len(df)} ({df['prefecture'].nunique()}都道府県)")

    # 重複チェック
    dupes = df.groupby(["year", "prefecture"]).size()
    dupes = dupes[dupes > 1]
    if len(dupes) > 0:
        print(f"  ⚠ 重複: {len(dupes)}件")

    return df


def parse_estat_production_openpyxl(path, year):
    """openpyxl で .xlsx ファイルから製成数量を抽出。"""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)

    # 製成数量シートを探す
    target_sheet = None
    for sn in wb.sheetnames:
        if "都道府県" in sn and "製成" in sn:
            target_sheet = sn
            break
    if not target_sheet:
        print(f"  ⚠ 製成数量シートが見つかりません: {path.name}")
        return []

    ws = wb[target_sheet]
    records = []
    for row_idx in range(1, ws.max_row + 1):
        c2 = ws.cell(row=row_idx, column=2).value
        pref = normalize_pref_name(str(c2)) if c2 else None
        if not pref:
            continue

        val = ws.cell(row=row_idx, column=4).value
        if val is None or val == "-" or val == "X":
            quantity = None
        else:
            try:
                quantity = float(val)
            except (ValueError, TypeError):
                quantity = None

        records.append({
            "year": year,
            "prefecture": pref,
            "production_quantity_kl": quantity,
        })

    return records


def parse_estat_production_xlrd(path, year):
    """xlrd で古い .xls ファイルから製成数量を抽出。"""
    import xlrd
    wb = xlrd.open_workbook(str(path))

    target_sheet = None
    for sn in wb.sheet_names():
        if "都道府県" in sn and "製成" in sn:
            target_sheet = sn
            break
    if not target_sheet:
        print(f"  ⚠ 製成数量シートが見つかりません: {path.name}")
        return []

    ws = wb.sheet_by_name(target_sheet)
    records = []
    for row_idx in range(ws.nrows):
        c2 = ws.cell_value(row_idx, 1) if ws.ncols > 1 else ""
        pref = normalize_pref_name(str(c2)) if c2 else None
        if not pref:
            continue

        val = ws.cell_value(row_idx, 3) if ws.ncols > 3 else None
        if val is None or val == "-" or val == "X" or val == "":
            quantity = None
        else:
            try:
                quantity = float(val)
            except (ValueError, TypeError):
                quantity = None

        records.append({
            "year": year,
            "prefecture": pref,
            "production_quantity_kl": quantity,
        })

    return records


def parse_estat_all():
    """e-Stat統計年報の全年度Excelから製成数量パネルを構築。"""
    print("\n=== e-Stat 統計年報 製成数量 パース ===")

    nenpo_dir = RAW_DIR / "nenpo"
    all_records = []

    for path in sorted(nenpo_dir.glob("estat_*_0802.*")):
        match = re.search(r"estat_(\d{4})_0802", path.name)
        if not match:
            continue
        year = int(match.group(1))

        print(f"  {year}年度: {path.name}")

        # ファイル形式に応じてパーサーを選択
        try:
            records = parse_estat_production_openpyxl(path, year)
        except Exception:
            try:
                records = parse_estat_production_xlrd(path, year)
            except Exception as e:
                print(f"    ✗ パース失敗: {e}")
                continue

        print(f"    → {len(records)}都道府県")
        all_records.extend(records)

    df = pd.DataFrame(all_records)
    print(f"  合計レコード数: {len(df)}")
    return df


def main():
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    # 1. 時系列Excel（販売数量）
    df_sales = parse_jikeiretsu()
    sales_path = PROCESSED_DIR / "sales_by_pref.csv"
    df_sales.to_csv(sales_path, index=False, encoding="utf-8-sig")
    print(f"  → 保存: {sales_path.name}")

    # 2. e-Stat製成数量
    df_prod = parse_estat_all()
    prod_path = PROCESSED_DIR / "production_by_pref_estat.csv"
    df_prod.to_csv(prod_path, index=False, encoding="utf-8-sig")
    print(f"  → 保存: {prod_path.name}")

    # サマリー
    print("\n=== サマリー ===")
    print(f"販売数量: {df_sales['prefecture'].nunique()}都道府県, "
          f"{df_sales['year'].min()}-{df_sales['year'].max()}年, "
          f"{len(df_sales)}レコード")
    print(f"製成数量: {df_prod['prefecture'].nunique()}都道府県, "
          f"{df_prod['year'].min()}-{df_prod['year'].max()}年, "
          f"{len(df_prod)}レコード")

    # 兵庫・新潟のスポットチェック
    print("\n--- スポットチェック: 販売数量 ---")
    for pref in ["兵庫", "新潟", "秋田"]:
        subset = df_sales[df_sales["prefecture"] == pref].sort_values("year")
        if len(subset) > 0:
            first = subset.iloc[0]
            last = subset.iloc[-1]
            print(f"  {pref}: {int(first['year'])}年={first['sales_quantity_kl']:.0f}kL → "
                  f"{int(last['year'])}年={last['sales_quantity_kl']:.0f}kL")

    print("\n--- スポットチェック: 製成数量 ---")
    for pref in ["兵庫", "新潟", "秋田"]:
        subset = df_prod[df_prod["prefecture"] == pref].sort_values("year")
        if len(subset) > 0:
            for _, row in subset.iterrows():
                q = row["production_quantity_kl"]
                q_str = f"{q:.0f}kL" if q is not None and pd.notna(q) else "N/A"
                print(f"  {pref} {int(row['year'])}年: {q_str}")


if __name__ == "__main__":
    main()
