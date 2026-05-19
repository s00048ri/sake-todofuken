"""
人口データパーサー

総務省統計局のExcelファイルから都道府県別人口を抽出し、
年次データに線形補間してCSVに出力する。

出典:
- 統計年鑑 第72回 (https://www.stat.go.jp/data/nenkan/72nenkan/)
  - y720203000.xlsx: 都道府県別人口 1920-2021 (5年おき + 2020/2021)
  - y720207000.xlsx: 都道府県別年齢3区分 2020-2021
- 人口推計 (https://www.stat.go.jp/data/jinsui/{YYYY}np/)
  - 05k{YYYY}-2.xlsx: 都道府県別人口 (年次, 2021-2024)
  - 05k{YYYY}-3.xlsx: 都道府県別年齢3区分 (年次, 2021-2024)

単位: 1,000 人（出典に合わせる）
"""

import sys
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = BASE_DIR / "data" / "raw" / "population"
PROCESSED_DIR = BASE_DIR / "data" / "processed"

sys.path.insert(0, str(BASE_DIR))
from scripts.parse_excel import PREFECTURES, normalize_pref_name


def parse_long_term_total():
    """y720203000.xlsx から都道府県別総人口の長期データを取得。

    各列 = 1年（5年おき）, 行 = 都道府県
    返り値: { year: { prefecture: population (千人) } }
    """
    path = RAW_DIR / "y720203000.xlsx"
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 列ヘッダから年を抽出（row 6 に "Population YYYY"）
    year_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=6, column=c).value
        if not v:
            continue
        s = str(v).strip()
        if s.startswith("Population"):
            try:
                year = int(s.replace("Population", "").strip())
                year_cols[year] = c
            except ValueError:
                continue

    print(f"  長期データ年度: {sorted(year_cols.keys())}")

    # 都道府県データを取得 (col 1 = 番号, col 2 = 都道府県名, col 3 = English)
    result = {y: {} for y in year_cols}
    for r in range(1, ws.max_row + 1):
        pref_raw = ws.cell(row=r, column=2).value
        if not pref_raw:
            continue
        pref = normalize_pref_name(str(pref_raw))
        if not pref:
            continue
        for year, col in year_cols.items():
            v = ws.cell(row=r, column=col).value
            if v is not None and isinstance(v, (int, float)):
                result[year][pref] = float(v)
    return result


def parse_jinsui_total(filepath):
    """jinsui の 05k{YYYY}-2.xlsx から都道府県別総人口を取得。

    年次データ。シート構造:
    - 列5 = 男女計 総人口
    - 行13以降 = 都道府県（北海道, 青森県, ...）
    - 単位: 1,000 人
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb[wb.sheetnames[0]]

    result = {}
    for r in range(1, ws.max_row + 1):
        pref_raw = ws.cell(row=r, column=3).value
        if not pref_raw:
            continue
        pref = normalize_pref_name(str(pref_raw))
        if not pref:
            continue
        total = ws.cell(row=r, column=5).value
        if total is not None and isinstance(total, (int, float)):
            result[pref] = float(total)
    return result


def parse_age_long_term():
    """y720207000.xlsx から都道府県別年齢3区分（2020, 2021）を取得。

    列構成:
    - col 3-6: 2020年 (総数, 0-14, 15-64, 65+)
    - col 8-11: 2021年
    """
    path = RAW_DIR / "y720207000.xlsx"
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    result = {2020: {}, 2021: {}}
    for r in range(1, ws.max_row + 1):
        pref_raw = ws.cell(row=r, column=1).value
        if not pref_raw:
            continue
        pref = normalize_pref_name(str(pref_raw))
        if not pref:
            continue
        for year, total_col, age0_col in [(2020, 3, 4), (2021, 8, 9)]:
            total = ws.cell(row=r, column=total_col).value
            age0 = ws.cell(row=r, column=age0_col).value
            if total is not None and age0 is not None:
                result[year][pref] = {
                    "total": float(total),
                    "age_0_14": float(age0),
                    "age_15_plus": float(total) - float(age0),
                }
    return result


def parse_jinsui_age(filepath):
    """jinsui の 05k{YYYY}-3.xlsx から都道府県別年齢3区分を取得。

    シート構造:
    - col 3 = 都道府県名
    - col 5 = 男女計 15歳未満
    - col 6 = 男女計 15-64歳
    - col 7 = 男女計 65歳以上
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb[wb.sheetnames[0]]

    result = {}
    for r in range(1, ws.max_row + 1):
        pref_raw = ws.cell(row=r, column=3).value
        if not pref_raw:
            continue
        pref = normalize_pref_name(str(pref_raw))
        if not pref:
            continue
        age0 = ws.cell(row=r, column=5).value
        age15_64 = ws.cell(row=r, column=6).value
        age65 = ws.cell(row=r, column=7).value
        if all(v is not None and isinstance(v, (int, float)) for v in [age0, age15_64, age65]):
            total = float(age0) + float(age15_64) + float(age65)
            result[pref] = {
                "total": total,
                "age_0_14": float(age0),
                "age_15_plus": float(age15_64) + float(age65),
            }
    return result


def linear_interpolate_annual(data_by_year, start_year=1963, end_year=2024):
    """5年おきデータを年次データに線形補間する。

    Args:
        data_by_year: {year: {prefecture: value}}
        start_year, end_year: 補間範囲

    Returns:
        DataFrame columns: year, prefecture, population
    """
    sorted_years = sorted(data_by_year.keys())
    if not sorted_years:
        return pd.DataFrame()

    # 全都道府県集合
    all_prefs = set()
    for y in sorted_years:
        all_prefs.update(data_by_year[y].keys())

    records = []
    for pref in all_prefs:
        # 各都道府県の (year, value) ペアを集める
        pairs = []
        for y in sorted_years:
            v = data_by_year[y].get(pref)
            if v is not None:
                pairs.append((y, v))

        if len(pairs) < 2:
            continue

        # start_year ~ end_year で年次補間
        for target_year in range(start_year, end_year + 1):
            # 範囲外チェック
            if target_year < pairs[0][0] or target_year > pairs[-1][0]:
                continue  # 補外しない

            # 該当年が直接データにあるか
            direct = next((v for y, v in pairs if y == target_year), None)
            if direct is not None:
                records.append({"year": target_year, "prefecture": pref, "population": direct})
                continue

            # 補間: target_yearを挟む2点を探す
            before = None
            after = None
            for y, v in pairs:
                if y <= target_year:
                    before = (y, v)
                if y >= target_year and after is None:
                    after = (y, v)

            if before and after and before[0] != after[0]:
                t = (target_year - before[0]) / (after[0] - before[0])
                value = before[1] + (after[1] - before[1]) * t
                records.append({"year": target_year, "prefecture": pref, "population": value})

    return pd.DataFrame(records)


def main():
    print("=== 都道府県別人口データ構築 ===\n")
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    # 1. 総人口（長期）
    print("[1/4] 長期総人口データ（統計年鑑 y720203000）...")
    long_total = parse_long_term_total()

    # 2. jinsui各年版で2022-2024を補完
    print("[2/4] 人口推計 (2022-2024) で総人口を補完...")
    for year in [2022, 2023, 2024]:
        f = RAW_DIR / f"05k{year}-2.xlsx"
        if f.exists():
            jinsui = parse_jinsui_total(f)
            if jinsui:
                long_total[year] = jinsui
                print(f"  {year}: {len(jinsui)}都道府県")

    # 線形補間で年次化（1963-2024）
    df_total = linear_interpolate_annual(long_total, 1963, 2024)
    df_total = df_total.rename(columns={"population": "population_total_k"})

    # 3. 年齢3区分（2020-2024）
    print("\n[3/4] 年齢3区分（2020-2024）...")
    age_data = parse_age_long_term()  # 2020-2021
    for year in [2022, 2023, 2024]:
        f = RAW_DIR / f"05k{year}-3.xlsx"
        if f.exists():
            jinsui_age = parse_jinsui_age(f)
            if jinsui_age:
                age_data[year] = jinsui_age
                print(f"  {year}: {len(jinsui_age)}都道府県")

    # DataFrame化
    age_records = []
    for year, prefs in age_data.items():
        for pref, vals in prefs.items():
            age_records.append({
                "year": year,
                "prefecture": pref,
                "age_0_14_k": vals["age_0_14"],
                "age_15_plus_k": vals["age_15_plus"],
            })
    df_age = pd.DataFrame(age_records)

    # 統合
    print("\n[4/4] 統合・保存...")
    df = df_total.merge(df_age, on=["year", "prefecture"], how="left")
    df = df.sort_values(["year", "prefecture"])

    out_path = PROCESSED_DIR / "population_by_pref.csv"
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"  → {out_path.name} ({len(df)}件)")

    # サマリー
    print("\n=== サマリー ===")
    print(f"年度範囲: {df['year'].min()}-{df['year'].max()}")
    print(f"都道府県数: {df['prefecture'].nunique()}")
    print(f"総人口: 全期間 ({df['population_total_k'].notna().sum()}件)")
    print(f"15歳以上: 2020-2024 ({df['age_15_plus_k'].notna().sum()}件)")

    print("\n--- スポットチェック 兵庫 ---")
    hyogo = df[df["prefecture"] == "兵庫"].sort_values("year")
    for y in [1965, 1973, 1985, 1995, 2010, 2020, 2024]:
        row = hyogo[hyogo["year"] == y]
        if len(row) > 0:
            r = row.iloc[0]
            total = r["population_total_k"]
            age15 = r["age_15_plus_k"]
            age15_str = f"{age15:,.0f}千" if pd.notna(age15) else "—"
            print(f"  {y}: 総人口={total:,.0f}千, 15歳以上={age15_str}")

    print("\n--- スポットチェック 新潟 ---")
    niigata = df[df["prefecture"] == "新潟"].sort_values("year")
    for y in [1965, 1985, 2024]:
        row = niigata[niigata["year"] == y]
        if len(row) > 0:
            r = row.iloc[0]
            print(f"  {y}: 総人口={r['population_total_k']:,.0f}千人")


if __name__ == "__main__":
    main()
